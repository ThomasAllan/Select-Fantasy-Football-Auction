"""
Import 2023-24 manager selections from FOOTY2023.xlsx into manager_selections.csv.

Usage:
    uv run python scripts/import_selections_2023_24.py [--dry-run]
"""

import argparse
import unicodedata
from pathlib import Path

import openpyxl
import pandas as pd

DATA_DIR = Path("data")
XLSX_PATH = Path("historic_selections/FOOTY2023.xlsx")
SEASON_ID = "2023-24"

MANAGER_SHEETS = [
    "Thomas Allan", "Pete Bezant", "Rory Canham", "Andrea Chapman", "Steve Fidler",
    "Andy Fowkes", "Tom Fowkes", "Steve Gale", "Pam Hart", "Tim Hart",
    "Mick Jones", "Ken Maggs", "Niall Mcloughlin", "Gary Speechley",
    "Jamie Wright", "Neil Wright",
]

TEAM_ALIASES: dict[str, str] = {
    "arsenal": "Arsenal",
    "aston villa": "Aston Villa",
    "bournemouth": "Bournemouth",
    "brentford": "Brentford",
    "brentord": "Brentford",
    "brighton ha": "Brighton",
    "brighton hove albion": "Brighton",
    "brighton": "Brighton",
    "burnley": "Burnley",
    "chelsea": "Chelsea",
    "crystal palace": "Crystal Palace",
    "c.palace": "Crystal Palace",
    "everton": "Everton",
    "fulham": "Fulham",
    "liverpool": "Liverpool",
    "luton town": "Luton",
    "luton": "Luton",
    "man city": "Man City",
    "manchester city": "Man City",
    "man united": "Man Utd",
    "man utd": "Man Utd",
    "manchester united": "Man Utd",
    "newcastle united": "Newcastle",
    "newcastle utd": "Newcastle",
    "newcastle": "Newcastle",
    "notts forest": "Nott'm Forest",
    "nottingham forest": "Nott'm Forest",
    "sheffield utd": "Sheffield Utd",
    "sheffield united": "Sheffield Utd",
    "spurs": "Spurs",
    "tottenham hotspur": "Spurs",
    "tottenham": "Spurs",
    "west ham": "West Ham",
    "west ham united": "West Ham",
    "west ham utd": "West Ham",
    "wolves": "Wolves",
    "wolverhampton wanderers": "Wolves",
}

# (norm_surname, norm_canonical_team) -> player_code
PLAYER_OVERRIDES: dict[tuple[str, str], str] = {
    ("jiminez", "fulham"): "2023-24-player-558",        # Raúl Jiménez (typo)
    ("vandijk", "liverpool"): "2023-24-player-313",     # Virgil van Dijk (friendly="Virgil")
    ("icasmeiro", "manutd"): "2023-24-player-376",      # Casemiro (wrong initial, no dot)
    ("schlup", "crystalpalace"): "2023-24-player-240",  # Schlupp (typo, one p)
    ("guimares", "newcastle"): "2023-24-player-406",    # Bruno Guimarães (typo)
    ("wellbeck", "brighton"): "2023-24-player-154",     # Welbeck (extra l)
    ("mcneill", "everton"): "2023-24-player-259",       # McNeil (extra l)
    ("periera", "fulham"): "2023-24-player-267",        # Andreas Pereira (typo)
    ("nunez", "liverpool"): "2023-24-player-293",       # Darwin Núñez (friendly="Darwin")
    ("sonheungmin", "spurs"): "2023-24-player-516",     # Son Heung-min (no dot prefix)
    ("dias", "mancity"): "2023-24-player-350",          # Rúben Dias (friendly="Rúben")
    ("chukuwemeka", "chelsea"): "2023-24-player-196",   # Chukwuemeka (u/w transposed)
    ("torres", "astonvilla"): "2023-24-player-584",     # Pau Torres (friendly="Pau")
    ("szobaszlai", "liverpool"): "2023-24-player-309",  # Szoboszlai (typo)
    ("alli", "everton"): "2023-24-player-248",          # Dele Alli (friendly="Dele")
    ("isoloman", "spurs"): "2023-24-player-583",        # Solomon (no dot, wrong format)
    ("hojlund", "manutd"): "2023-24-player-617",        # Højlund (ø not handled)
    ("odegaard", "arsenal"): "2023-24-player-14",       # Ødegaard (ø not handled)
    ("grob", "brighton"): "2023-24-player-134",         # Groß / Gross (typo)
    ("manuel", "burnley"): "2023-24-player-159",        # Manuel Benson (first name used)
    ("ammehodzic", "sheffieldutd"): "2023-24-player-471",  # Ahmedhodzic (typo)
    ("isasi", "chelsea"): "2023-24-player-611",         # Disasi (missing D)
    ("pjdasilva", "brentford"): "2023-24-player-99",   # Dasilva (no dot format)
    ("jota", "liverpool"): "2023-24-player-294",        # Diogo J. (friendly="Diogo J.")
    ("silva", "mancity"): "2023-24-player-344",         # Bernardo Silva (friendly="Bernardo")
    ("sarr", "chelsea"): "2023-24-player-513",          # Pape Matar Sarr (listed as Chelsea in Excel, actually Spurs)
}


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return "".join(c for c in s.lower() if c.isalpha())


def normalize_team(name: str) -> str:
    return TEAM_ALIASES.get(name.strip().lower(), name.strip())


def extract_surname(excel_name: str) -> str:
    name = excel_name.strip()
    if "." in name:
        return name.split(".", 1)[1].strip()
    return name


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    players_df = pd.read_csv(DATA_DIR / "players.csv", dtype=str).fillna("")
    season_p = players_df[players_df["season"] == SEASON_ID]
    outfield = season_p[season_p["type"] == "player"]
    teams_df = season_p[season_p["type"] == "team"]

    team_id_lookup: dict[str, str] = {row["full_name"]: row["element_id"] for _, row in teams_df.iterrows()}
    tid_to_name: dict[str, str] = {row["element_id"]: row["full_name"] for _, row in teams_df.iterrows()}

    player_lookup: dict[tuple[str, str], str] = {}
    surname_all: dict[str, list[str]] = {}

    for _, row in outfield.iterrows():
        tname = tid_to_name.get(row["team_id"], "")
        fname = _norm(row["friendly_name"])
        nt = _norm(tname)
        player_lookup[(fname, nt)] = row["code"]
        if "." in row["friendly_name"]:
            suffix = row["friendly_name"].split(".", 1)[1].strip()
            player_lookup[(_norm(suffix), nt)] = row["code"]
        parts = row["friendly_name"].split()
        if len(parts) > 1:
            player_lookup[(_norm(parts[-1]), nt)] = row["code"]
        surname_all.setdefault(fname, []).append(row["code"])

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    all_rows: list[dict] = []
    unmatched: list[str] = []
    POSITION_ORDER = ["GK", "DEF", "DEF", "DEF", "MID", "MID", "MID", "MID", "FWD", "FWD", "FWD"]

    SHEET_NAME_OVERRIDES = {"Thomas Allan": "Tom Allan"}

    for manager in MANAGER_SHEETS:
        sheet_name = SHEET_NAME_OVERRIDES.get(manager, manager)
        if sheet_name not in wb.sheetnames:
            print(f"WARN: sheet '{sheet_name}' not found")
            continue

        ws = wb[sheet_name]
        sheet_rows = list(ws.iter_rows(min_row=8, max_row=18, values_only=True))
        print(f"\n=== {manager} ===")

        for raw_row, position in zip(sheet_rows, POSITION_ORDER):
            col_a = str(raw_row[0] or "").strip()
            col_b = str(raw_row[1] or "").strip()
            col_c = raw_row[2]
            cost = str(int(col_c)) if isinstance(col_c, (int, float)) else str(col_c or "")

            if position == "GK":
                canonical = normalize_team(col_a)
                team_id = team_id_lookup.get(canonical, "")
                if not team_id:
                    print(f"  UNMATCHED GK: '{col_a}' -> '{canonical}'")
                    unmatched.append(f"{manager} | GK | {col_a}")
                    continue
                player_code = f"{SEASON_ID}-team-{team_id}"
                print(f"  GK  {col_a:25s} -> {player_code}  {cost}")
            else:
                surname = extract_surname(col_a)
                canonical_team = normalize_team(col_b)
                norm_s = _norm(surname)
                norm_t = _norm(canonical_team)

                player_code = (
                    PLAYER_OVERRIDES.get((norm_s, norm_t), "")
                    or player_lookup.get((norm_s, norm_t), "")
                )
                if not player_code:
                    matches = surname_all.get(norm_s, [])
                    if len(matches) == 1:
                        player_code = matches[0]

                if not player_code:
                    print(f"  UNMATCHED {position}: '{col_a}' @ '{col_b}'")
                    unmatched.append(f"{manager} | {position} | {col_a} @ {col_b}")
                    continue

                print(f"  {position}  {col_a:25s} {col_b:20s} -> {player_code}  {cost}")

            all_rows.append({
                "player_code": player_code,
                "season_id": SEASON_ID,
                "manager_name": manager,
                "position": position,
                "cost": cost,
                "gw_from": "1",
                "gw_to": "38",
            })

    print(f"\n{'='*60}")
    print(f"Matched:   {len(all_rows)}")
    print(f"Unmatched: {len(unmatched)}")
    if unmatched:
        print("\nUnmatched:")
        for u in unmatched:
            print(f"  {u}")

    if not args.dry_run and all_rows:
        new_df = pd.DataFrame(all_rows)
        sel_path = DATA_DIR / "manager_selections.csv"
        if sel_path.exists() and sel_path.stat().st_size > 0:
            existing = pd.read_csv(sel_path, dtype=str).fillna("")
            existing = existing[existing["season_id"] != SEASON_ID]
            result = pd.concat([existing, new_df], ignore_index=True)
        else:
            result = new_df
        result.to_csv(sel_path, index=False)
        print(f"\nWritten {len(new_df)} rows to manager_selections.csv")
    elif args.dry_run:
        print("\nDry run - no changes written.")


if __name__ == "__main__":
    main()
