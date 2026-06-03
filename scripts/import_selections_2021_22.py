"""
Import 2021-22 manager selections from FOOTY2021.xlsx into manager_selections.csv.

This season uses full player names (not "I.Surname" format).
Tom Allan did not participate this season.
Karl Allen and Jamie Blunt participated (historical managers).

Usage:
    uv run python scripts/import_selections_2021_22.py [--dry-run]
"""

import argparse
import unicodedata
from pathlib import Path

import openpyxl
import pandas as pd

DATA_DIR = Path("data")
XLSX_PATH = Path("historic_selections/FOOTY2021.xlsx")
SEASON_ID = "2021-22"

MANAGER_SHEETS = [
    "Karl Allen", "Jamie Blunt", "Andrea Chapman", "Andy Fowkes", "Tom Fowkes",
    "Steve Gale", "Pam Hart", "Tim Hart", "Mick Jones", "Ken Maggs",
    "Gary Speechley", "Kev Thulbourne", "Wendy Thulbourne", "Jamie Wright", "Neil Wright",
    # Tom Allan did not play this season
]

TEAM_ALIASES: dict[str, str] = {
    "arsenal": "Arsenal", "aston villa": "Aston Villa", "brentford": "Brentford",
    "brighton ha": "Brighton", "brighton": "Brighton", "burnley": "Burnley",
    "c.palace": "Crystal Palace", "chelsea": "Chelsea", "crystal palace": "Crystal Palace",
    "everton": "Everton", "leeds united": "Leeds", "leeds utd": "Leeds", "leeds": "Leeds",
    "leicester city": "Leicester", "leicester": "Leicester", "liverpool": "Liverpool",
    "man city": "Man City", "manchester city": "Man City",
    "man utd": "Man Utd", "man united": "Man Utd", "manchester united": "Man Utd",
    "newcastle utd": "Newcastle", "newcastle": "Newcastle",
    "norwich city": "Norwich", "norwich": "Norwich",
    "southampton": "Southampton", "spurs": "Spurs", "tottenham": "Spurs",
    "watford": "Watford", "west ham utd": "West Ham", "west ham": "West Ham",
    "wolves": "Wolves", "wolverhampton wanderers": "Wolves",
}

# (norm_display_name, norm_canonical_team) -> player_code
# Needed because this season uses full display names which often don't match
# vaastav's official full_name or friendly_name exactly.
PLAYER_OVERRIDES: dict[tuple[str, str], str] = {
    ("brunofernandes", "manutd"): "2021-22-player-277",         # full_name is much longer
    ("raphina", "leeds"): "2021-22-player-196",                 # typo: Raphina vs Raphinha
    ("hectorfirpo", "leeds"): "2021-22-player-463",             # accent stripped ok, Firpo
    ("trentalexarnold", "liverpool"): "2021-22-player-237",     # abbreviated vs Alexander-Arnold
    ("delealli", "everton"): "2021-22-player-363",              # friendly="Dele", full_name longer
    ("pedroneto", "wolves"): "2021-22-player-441",              # full_name has middle name
    ("heldercosta", "wolves"): "2021-22-player-192",            # Excel wrong team (Leeds), link correct player
    ("pcoutinho", "astonvilla"): "2021-22-player-681",          # P. prefix
    ("madssørensen", "brentford"): "2021-22-player-90",    # ø not decomposed by NFD
    ("nelsonsemedo", "wolves"): "2021-22-player-437",           # full_name much longer
    ("franciscotrincao", "wolves"): "2021-22-player-461",       # ã stripped to a ok
    ("rubendias", "mancity"): "2021-22-player-262",             # full_name much longer
    ("emilianobuendia", "astonvilla"): "2021-22-player-43",     # í stripped to i ok
    ("connorgallagher", "crystalpalace"): "2021-22-player-144", # safety override
    ("gabrieljesus", "mancity"): "2021-22-player-263",          # full_name has middle names
    ("albertslokonga", "arsenal"): "2021-22-player-478",        # "Albert S Lokonga" abbreviated
    ("wweghurst", "burnley"): "2021-22-player-700",             # W. prefix + Weghurst typo
    ("alexoxlchamberlain", "liverpool"): "2021-22-player-227",  # abbreviated Oxlade-Chamberlain
    ("rubenneves", "wolves"): "2021-22-player-436",             # full_name much longer
}


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return "".join(c for c in s.lower() if c.isalpha())


def normalize_team(name: str) -> str:
    return TEAM_ALIASES.get(name.strip().lower(), name.strip())


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    players_df = pd.read_csv(DATA_DIR / "players.csv", dtype=str).fillna("")
    season_p = players_df[players_df["season"] == SEASON_ID]
    outfield = season_p[season_p["type"] == "player"]
    teams_df = season_p[season_p["type"] == "team"]

    team_id_lookup: dict[str, str] = {row["full_name"]: row["element_id"] for _, row in teams_df.iterrows()}
    tid_to_name: dict[str, str] = {row["element_id"]: row["full_name"] for _, row in teams_df.iterrows()}

    full_lookup: dict[tuple[str, str], str] = {}
    friendly_lookup: dict[tuple[str, str], str] = {}
    all_full: dict[str, list[str]] = {}
    all_friendly: dict[str, list[str]] = {}

    for _, row in outfield.iterrows():
        tname = tid_to_name.get(row["team_id"], "")
        nt = _norm(tname)
        fn = _norm(row["full_name"])
        wb_n = _norm(row["friendly_name"])
        full_lookup[(fn, nt)] = row["code"]
        friendly_lookup[(wb_n, nt)] = row["code"]
        parts = row["friendly_name"].split()
        if len(parts) > 1:
            friendly_lookup[(_norm(parts[-1]), nt)] = row["code"]
        all_full.setdefault(fn, []).append(row["code"])
        all_friendly.setdefault(wb_n, []).append(row["code"])

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    all_rows: list[dict] = []
    unmatched: list[str] = []
    POSITION_ORDER = ["GK", "DEF", "DEF", "DEF", "MID", "MID", "MID", "MID", "MID", "FWD", "FWD"]  # 3-5-2

    for manager in MANAGER_SHEETS:
        if manager not in wb.sheetnames:
            print(f"WARN: sheet '{manager}' not found")
            continue

        ws = wb[manager]
        sheet_rows = list(ws.iter_rows(min_row=8, max_row=18, values_only=True))
        print(f"\n=== {manager} ===")

        for raw_row, position in zip(sheet_rows, POSITION_ORDER):
            col_a = str(raw_row[0] or "").strip()
            col_b = str(raw_row[1] or "").strip()
            col_c = raw_row[2]
            cost = str(int(col_c)) if isinstance(col_c, (int, float)) else str(col_c or "")

            if position == "GK":
                canonical = normalize_team(col_a)
                team_id = team_id_lookup.get(canonical, "")
                if not team_id:
                    print(f"  UNMATCHED GK: '{col_a}'")
                    unmatched.append(f"{manager} | GK | {col_a}")
                    continue
                player_code = f"{SEASON_ID}-team-{team_id}"
                print(f"  GK  {col_a:28s} -> {player_code}  {cost}")
            else:
                canonical_team = normalize_team(col_b)
                norm_t = _norm(canonical_team)
                norm_full = _norm(col_a)

                player_code = (
                    PLAYER_OVERRIDES.get((norm_full, norm_t), "")
                    or full_lookup.get((norm_full, norm_t), "")
                    or friendly_lookup.get((norm_full, norm_t), "")
                )
                if not player_code:
                    hits = all_full.get(norm_full, [])
                    if len(hits) == 1:
                        player_code = hits[0]
                if not player_code:
                    hits = all_friendly.get(norm_full, [])
                    if len(hits) == 1:
                        player_code = hits[0]

                if not player_code:
                    print(f"  UNMATCHED {position}: '{col_a}' @ '{col_b}'")
                    unmatched.append(f"{manager} | {position} | {col_a} @ {col_b}")
                    continue

                print(f"  {position}  {col_a:28s} {col_b:16s} -> {player_code}  {cost}")

            all_rows.append({
                "player_code": player_code,
                "season_id": SEASON_ID,
                "manager_name": manager,
                "position": position,
                "cost": cost,
                "gw_from": "1",
                "gw_to": "38",
            })

    print(f"\n{'='*60}")
    print(f"Matched:   {len(all_rows)}")
    print(f"Unmatched: {len(unmatched)}")
    if unmatched:
        print("\nUnmatched:")
        for u in unmatched:
            print(f"  {u}")

    if not args.dry_run and all_rows:
        new_df = pd.DataFrame(all_rows)
        sel_path = DATA_DIR / "manager_selections.csv"
        if sel_path.exists() and sel_path.stat().st_size > 0:
            existing = pd.read_csv(sel_path, dtype=str).fillna("")
            existing = existing[existing["season_id"] != SEASON_ID]
            result = pd.concat([existing, new_df], ignore_index=True)
        else:
            result = new_df
        result.to_csv(sel_path, index=False)
        print(f"\nWritten {len(new_df)} rows to manager_selections.csv")
    elif args.dry_run:
        print("\nDry run - no changes written.")


if __name__ == "__main__":
    main()
