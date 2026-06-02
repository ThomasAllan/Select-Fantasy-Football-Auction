"""
Import 2024-25 manager selections from FOOTY2024.xlsx into manager_selections.csv.

Each manager sheet has:
  Row 8:      GK team   (col A = team name, col B = None, col C = cost)
  Rows 9-11:  DEF players (col A = "I.Surname", col B = team, col C = cost)
  Rows 12-15: MID players
  Rows 16-18: FWD players

Usage:
    uv run python scripts/import_selections_2024_25.py [--dry-run]
"""

import argparse
import unicodedata
from pathlib import Path

import openpyxl
import pandas as pd

DATA_DIR = Path("data")
XLSX_PATH = Path("historic_selections/FOOTY2024.xlsx")
SEASON_ID = "2024-25"

MANAGER_SHEETS = [
    "Tom Allan", "Rory Canham", "Andrea Chapman", "Steve Fidler",
    "Andy Fowkes", "Tom Fowkes", "Steve Gale", "Pam Hart", "Tim Hart",
    "Mick Jones", "Ken Maggs", "Niall Mcloughlin", "Gary Speechley",
    "Jamie Wright", "Neil Wright",
]

# (norm_surname, norm_canonical_team) -> player_code
# For cases where Excel name can't auto-match vaastav friendly_name
PLAYER_OVERRIDES: dict[tuple[str, str], str] = {
    ("guimares", "newcastle"): "2024-25-player-394",   # Bruno G. (vaastav abbreviates)
    ("anderson", "fulham"): "2024-25-player-191",      # Joachim Andersen (vaastav "Andersen" at Fulham, not Elliot Anderson at Newcastle)
    ("burns", "newcastle"): "2024-25-player-395",      # Dan Burn (vaastav "Burn" no s, at Newcastle)
    ("silva", "mancity"): "2024-25-player-342",        # Bernardo Silva (not B.Silva player-753)
    ("vandijk", "liverpool"): "2024-25-player-339",    # Virgil van Dijk (friendly="Virgil")
    ("vvden", "spurs"): "2024-25-player-506",           # Van de Ven (typo VVDen)
    ("rodri", "mancity"): "2024-25-player-360",        # Rodrigo 'Rodri' (friendly="Rodrigo")
    ("heungmin", "spurs"): "2024-25-player-503",       # Son Heung-min (friendly="Son")
    ("odegaard", "arsenal"): "2024-25-player-13",      # Odegaard (ø → o not handled)
    ("torres", "astonvilla"): "2024-25-player-52",     # Pau Torres (friendly="Pau")
    ("kudos", "westham"): "2024-25-player-525",        # Kudus typo
    ("hojlund", "manutd"): "2024-25-player-375",       # Hojlund (ø → o not handled)
    ("jiminez", "fulham"): "2024-25-player-252",       # Jiménez (friendly="Raúl")
    ("nunez", "liverpool"): "2024-25-player-316",      # Darwin Núñez (friendly="Darwin")
    ("ouattara", "bournemouth"): "2024-25-player-74",  # O.Dango Ouattara
    ("sugarwara", "southampton"): "2024-25-player-474", # Sugawara typo
    ("neketiah", "crystalpalace"): "2024-25-player-11", # Nketiah typo
    ("periera", "fulham"): "2024-25-player-240",        # Pereira typo (Andreas)
    ("nkunu", "chelsea"): "2024-25-player-181",        # Nkunku typo
    ("jota", "liverpool"): "2024-25-player-317",       # Diogo J. (friendly="Diogo J.")
}

TEAM_ALIASES: dict[str, str] = {
    "arsenal": "Arsenal",
    "aston villa": "Aston Villa",
    "bournemouth": "Bournemouth",
    "brentford": "Brentford",
    "brighton ha": "Brighton",
    "brighton hove albion": "Brighton",
    "brighton": "Brighton",
    "chelsea": "Chelsea",
    "c.palace": "Crystal Palace",
    "crystal palace": "Crystal Palace",
    "everton": "Everton",
    "fulham": "Fulham",
    "ipswich town": "Ipswich",
    "ipswich": "Ipswich",
    "leicester city": "Leicester",
    "leicester": "Leicester",
    "liverpool": "Liverpool",
    "man city": "Man City",
    "manchester city": "Man City",
    "man united": "Man Utd",
    "man utd": "Man Utd",
    "manchester united": "Man Utd",
    "newcastle united": "Newcastle",
    "newcastle utd": "Newcastle",
    "newcastle": "Newcastle",
    "notts forest": "Nott'm Forest",
    "nottingham forest": "Nott'm Forest",
    "southampton": "Southampton",
    "spurs": "Spurs",
    "tottenham": "Spurs",
    "west ham utd": "West Ham",
    "west ham": "West Ham",
    "wolverhampton wanderers": "Wolves",
    "wolves": "Wolves",
}


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return "".join(c for c in s.lower() if c.isalpha())


def normalize_team(name: str) -> str:
    return TEAM_ALIASES.get(name.strip().lower(), name.strip())


def extract_surname(excel_name: str) -> str:
    """'B.Saka' -> 'Saka', 'K.De Bruyne' -> 'De Bruyne'"""
    name = excel_name.strip()
    if "." in name:
        return name.split(".", 1)[1].strip()
    return name


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    players_df = pd.read_csv(DATA_DIR / "players.csv", dtype=str).fillna("")
    season_p = players_df[players_df["season"] == SEASON_ID]
    outfield = season_p[season_p["type"] == "player"]
    teams_df = season_p[season_p["type"] == "team"]

    # team full_name -> element_id
    team_id_lookup: dict[str, str] = {row["full_name"]: row["element_id"] for _, row in teams_df.iterrows()}
    # team element_id -> full_name (for reverse lookup)
    tid_to_name: dict[str, str] = {row["element_id"]: row["full_name"] for _, row in teams_df.iterrows()}

    # Build (norm_surname, norm_team_full_name) -> player_code
    player_lookup: dict[tuple[str, str], str] = {}
    # Fallback: norm_surname -> list of codes (for unique-surname matches)
    surname_all: dict[str, list[str]] = {}

    for _, row in outfield.iterrows():
        tname = tid_to_name.get(row["team_id"], "")
        fname = _norm(row["friendly_name"])
        nt = _norm(tname)
        player_lookup[(fname, nt)] = row["code"]
        # Vaastav stores some players as "X.Surname" — also index just the surname
        if "." in row["friendly_name"]:
            suffix = row["friendly_name"].split(".", 1)[1].strip()
            player_lookup[(_norm(suffix), nt)] = row["code"]
        # Also index last word of multi-word friendly names (e.g. "Van de Ven" -> "ven")
        parts = row["friendly_name"].split()
        if len(parts) > 1:
            player_lookup[(_norm(parts[-1]), nt)] = row["code"]
        surname_all.setdefault(fname, []).append(row["code"])

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    all_rows: list[dict] = []
    unmatched: list[str] = []

    POSITION_ORDER = ["GK", "DEF", "DEF", "DEF", "MID", "MID", "MID", "MID", "FWD", "FWD", "FWD"]

    for manager in MANAGER_SHEETS:
        if manager not in wb.sheetnames:
            print(f"WARN: sheet '{manager}' not found")
            continue

        ws = wb[manager]
        sheet_rows = list(ws.iter_rows(min_row=8, max_row=18, values_only=True))
        if len(sheet_rows) != 11:
            print(f"WARN: {manager} — expected 11 rows, got {len(sheet_rows)}")
            continue

        print(f"\n=== {manager} ===")
        for raw_row, position in zip(sheet_rows, POSITION_ORDER):
            col_a = str(raw_row[0] or "").strip()
            col_b = str(raw_row[1] or "").strip()
            col_c = raw_row[2]
            cost = str(int(col_c)) if isinstance(col_c, (int, float)) else str(col_c or "")

            if position == "GK":
                canonical = normalize_team(col_a)
                team_id = team_id_lookup.get(canonical, "")
                if not team_id:
                    print(f"  UNMATCHED GK: '{col_a}' -> '{canonical}'")
                    unmatched.append(f"{manager} | GK | {col_a}")
                    continue
                player_code = f"{SEASON_ID}-team-{team_id}"
                print(f"  GK  {col_a:25s} -> {player_code}  £{cost}")
            else:
                surname = extract_surname(col_a)
                canonical_team = normalize_team(col_b)
                norm_s = _norm(surname)
                norm_t = _norm(canonical_team)

                player_code = (
                    PLAYER_OVERRIDES.get((norm_s, norm_t), "")
                    or player_lookup.get((norm_s, norm_t), "")
                )

                # Fallback: unique surname across whole season
                if not player_code:
                    matches = surname_all.get(norm_s, [])
                    if len(matches) == 1:
                        player_code = matches[0]

                if not player_code:
                    print(f"  UNMATCHED {position}: '{col_a}' @ '{col_b}' (surname='{surname}', team='{canonical_team}')")
                    unmatched.append(f"{manager} | {position} | {col_a} @ {col_b}")
                    continue

                print(f"  {position}  {col_a:25s} {col_b:20s} -> {player_code}  £{cost}")

            all_rows.append({
                "player_code": player_code,
                "season_id": SEASON_ID,
                "manager_name": manager,
                "position": position,
                "cost": cost,
                "gw_from": "1",
                "gw_to": "38",
            })

    print(f"\n{'='*60}")
    print(f"Matched:   {len(all_rows)}")
    print(f"Unmatched: {len(unmatched)}")
    if unmatched:
        print("\nUnmatched players (need manual fix):")
        for u in unmatched:
            print(f"  {u}")

    if not args.dry_run and all_rows:
        new_df = pd.DataFrame(all_rows)
        sel_path = DATA_DIR / "manager_selections.csv"
        if sel_path.exists() and sel_path.stat().st_size > 0:
            existing = pd.read_csv(sel_path, dtype=str).fillna("")
            existing = existing[existing["season_id"] != SEASON_ID]
            result = pd.concat([existing, new_df], ignore_index=True)
        else:
            result = new_df
        result.to_csv(sel_path, index=False)
        print(f"\nWritten {len(new_df)} rows to manager_selections.csv")
    elif args.dry_run:
        print("\nDry run — no changes written.")


if __name__ == "__main__":
    main()
