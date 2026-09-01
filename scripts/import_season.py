"""
Import manager selections from a FOOTY20XX.xlsx workbook for any season.

Designed for ongoing use at the start of each new season. Run sync-scores first
so that players.csv is populated for the new season, then run this script.

USAGE
-----
    # Drop this season's workbook into import/ then just run:
    uv run python scripts/import_season.py

    # first pass: preview + let it auto-accept confident guesses
    uv run python scripts/import_season.py --dry-run --auto

    # ...or point at a file explicitly / tweak the knobs:
    uv run python scripts/import_season.py FOOTY2026.xlsx --season 2026-27
    uv run python scripts/import_season.py --overrides import/fixes.csv
    uv run python scripts/import_season.py --shape 3-5-2 --budget 100 --strict

On a successful write, a workbook picked up from import/ and the default
import/fixes.csv are both deleted (the data now lives in manager_selections.csv).
An explicit workbook path or a custom --overrides path is never deleted, and
--dry-run never deletes anything.

EXCEL WORKBOOK LAYOUT (one sheet per manager)
---------------------------------------------
  Row 8:       GK team       (col A = team name,  col C = cost)
  Rows 9-18:   10 outfield   (col A = I.Surname,  col B = club, col C = cost)
  Row 19 col C: squad total  (cross-checked against the summed costs)

Cost can be a number (13), a float (13.0), or a string with £ prefix (£13).

SQUAD SHAPE
-----------
Rows 9-18 are split into DEF / MID / FWD by the season's shape. Set it in
SEASON_SHAPES below (e.g. "2026-27": (3, 5, 2)); otherwise DEFAULT_SHAPE (3-4-3)
is used. Override for one run with --shape 3-5-2. Getting this wrong mis-scores
every player, so the shape actually used is printed at the top of the run.

CHECKS (run before a real import; a --dry-run only reports them)
--------------------------------------------------------------
PROBLEMS — block a real import (exit non-zero):
  * a matched player whose FPL club != the club written on the sheet (col B).
    Catches the "unique surname" fallback grabbing the wrong player. To accept it
    (e.g. the player changed club after the auction), pin that player_code in the
    fixes CSV for that row — an --overrides match is treated as "checked", and
    the mismatch is just noted, not blocked. (Or fix col B in the workbook.)
  * a player or GK team picked by more than one manager (duplicate)
  * a squad that costs more than --budget (default 100)
Under --strict the warnings below also become blocking.

WARNINGS — shown, but don't block unless --strict:
  * a squad whose costs don't add up to the sheet's own row-19 total (typo)

Also printed: club mismatches accepted via the fixes CSV, managers who spent
under budget (both allowed), and a reminder to run sync-scores if seasons.csv
shows players.csv is more than a few days stale.

THE FIXES CSV (import/fixes.csv)
-------------------------------
When a run has anything to resolve — unmatched rows, or a PROBLEM (club mismatch
/ duplicate) — it writes them all to import/fixes.csv (or to your --overrides
FILE if you passed one that doesn't exist yet). A real run then exits non-zero; a
--dry-run just reports and exits 0. Columns:

    manager,position,excel_name,excel_team,player_code,note,suggestions

  * note        — why the row is here ("unmatched", "club? matched X ...",
                  "duplicate 2x — change one")
  * player_code — blank for an unmatched row; pre-filled with the current guess
                  for a club-mismatch / duplicate row. Leave it to accept that
                  guess, change it to correct the match.
  * suggestions — likeliest players and their codes; usually the first is right.

Fill / correct player_code, then re-run with --overrides import/fixes.csv. Repeat
until the run is clean. An existing fixes file is never overwritten, so edits
across several runs are safe. Only excel_name/excel_team/player_code are read
back; manager, position, note and suggestions are ignored.

For GK teams, excel_team is blank and player_code is a "{season}-team-{id}" code.
To look up codes by hand, open data/players.csv and filter by season + full_name.

SHEET NAME ALIASES
------------------
If a manager's sheet name differs from their canonical name (e.g. "Tom Allan"
vs "Thomas Allan"), add an entry to scripts/manager_names.py under
CANONICAL_NAMES. The script uses that mapping automatically.

SEASON AUTO-DETECTION
---------------------
Without --season, the script picks the most recent season that has player data
in players.csv. Run 'uv run sync-scores' first to populate it.
"""

import argparse
import difflib
import re
import sys
import unicodedata
from datetime import UTC, datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
import pandas as pd

from manager_names import canonicalize
from team_aliases import TEAM_ALIASES

DATA_DIR = Path("data")
IMPORT_DIR = Path("import")
SCRIPTS_DIR = Path(__file__).parent

# Outfield squad shape as (DEF, MID, FWD) — must total 10. Plus 1 GK = sheet rows 8-18.
DEFAULT_SHAPE = (3, 4, 3)
SEASON_SHAPES: dict[str, tuple[int, int, int]] = {
    "2026-27": (3, 5, 2),
}

DEFAULT_BUDGET = 100       # auction spend per manager; mismatches warn (error under --strict)
STALE_SYNC_DAYS = 3        # remind to re-run sync-scores if players.csv is older than this
AUTO_MIN_SCORE = 0.80      # --auto: min surname similarity to auto-accept a same-club guess

# Sheet names that are not manager sheets — skip these
SKIP_SHEETS = {
    "summary", "rules", "prizes", "totals", "points", "league", "league table",
    "leaderboard", "other", "template", "example", "instructions",
    "sheet1", "sheet2", "sheet3", "overview", "scores", "standings", "history",
}

# TEAM_ALIASES (any spelling a manager might type → canonical FPL full_name) lives
# in scripts/team_aliases.py. Add promoted clubs / new spellings there.


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _norm(s: str) -> str:
    """Lowercase, strip diacritics (é→e, ø→o, etc.), remove non-alpha chars."""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return "".join(c for c in s.lower() if c.isalpha())


def _clean_cost(val: object) -> str:
    if isinstance(val, (int, float)):
        return str(int(val))
    cleaned = re.sub(r"[£$,\s]", "", str(val or ""))
    try:
        return str(int(float(cleaned)))
    except ValueError:
        return "0"


def _extract_surname(excel_name: str) -> str:
    """
    Extract the surname from an Excel player name.
    'B.Saka' -> 'Saka'
    'K.De Bruyne' -> 'De Bruyne'
    'Haaland' -> 'Haaland'
    """
    name = excel_name.strip()
    if "." in name:
        return name.split(".", 1)[1].strip()
    return name


def _normalize_team(name: str) -> str | None:
    """Return canonical FPL full_name for a team alias, or None if unknown."""
    return TEAM_ALIASES.get(name.strip().lower())


def _parse_shape(text: str) -> tuple[int, int, int]:
    """Parse a '3-5-2' DEF-MID-FWD shape string."""
    try:
        d, m, f = (int(x) for x in text.split("-"))
    except ValueError:
        print(f"ERROR: --shape must look like 3-5-2 (DEF-MID-FWD), got {text!r}")
        sys.exit(1)
    return d, m, f


def _position_order(shape: tuple[int, int, int]) -> list[str]:
    """['GK', 'DEF', 'DEF', 'DEF', 'MID', ...] for sheet rows 8-18."""
    d, m, f = shape
    return ["GK"] + ["DEF"] * d + ["MID"] * m + ["FWD"] * f


# ---------------------------------------------------------------------------
# Build lookups from players.csv
# ---------------------------------------------------------------------------

def _build_lookups(players_df: pd.DataFrame, season_id: str) -> tuple[
    dict[str, str],              # team full_name -> element_id
    dict[str, str],              # team element_id -> full_name
    dict[tuple[str, str], str],  # (norm_name_variant, norm_team) -> player_code
    dict[str, list[str]],        # norm_friendly_name -> [player_codes]
]:
    season_p = players_df[players_df["season"] == season_id]
    teams_df = season_p[season_p["type"] == "team"]
    outfield = season_p[season_p["type"] == "player"]

    team_id_lookup: dict[str, str] = {}
    tid_to_name: dict[str, str] = {}
    for _, r in teams_df.iterrows():
        eid = str(r["element_id"])
        fname = str(r["full_name"])
        team_id_lookup[fname] = eid
        tid_to_name[eid] = fname

    player_lookup: dict[tuple[str, str], str] = {}
    surname_all: dict[str, list[str]] = {}

    for _, r in outfield.iterrows():
        tname = tid_to_name.get(str(r.get("team_id", "")), "")
        nt = _norm(tname)
        code = str(r["code"])
        fn = str(r.get("friendly_name", "") or "")
        full = str(r.get("full_name", "") or "")

        # Index by full FPL friendly_name (normalised)
        norm_fn = _norm(fn)
        player_lookup[(norm_fn, nt)] = code

        # If FPL uses "X.Surname" format, also index just the surname part
        if "." in fn:
            suffix = fn.split(".", 1)[1].strip()
            player_lookup[(_norm(suffix), nt)] = code

        # Last word of multi-word friendly_name (e.g. "Van de Ven" → "ven")
        parts = fn.split()
        if len(parts) > 1:
            player_lookup[(_norm(parts[-1]), nt)] = code
            # First word too (e.g. "Diogo J." → "diogo" helps match "D.Jota")
            player_lookup[(_norm(parts[0]), nt)] = code

        # FPL full_name last word (e.g. full_name="Darwin Núñez" → "nunez")
        fparts = full.split()
        if fparts:
            player_lookup[(_norm(fparts[-1]), nt)] = code
            if len(fparts) > 1:
                player_lookup[(_norm(fparts[0]), nt)] = code

        # surname_all: for unique-across-season fallback
        surname_all.setdefault(norm_fn, []).append(code)

    return team_id_lookup, tid_to_name, player_lookup, surname_all


# ---------------------------------------------------------------------------
# Player matching
# ---------------------------------------------------------------------------

def _match_player(
    excel_name: str,
    excel_team: str,
    player_lookup: dict[tuple[str, str], str],
    surname_all: dict[str, list[str]],
    overrides: dict[tuple[str, str], str],
) -> tuple[str, str]:
    """
    Try to find a player_code for an Excel name+team combination.
    Returns (player_code, method_description).
    Returns ("", "") if no match found.
    """
    surname = _extract_surname(excel_name)
    canonical_team = _normalize_team(excel_team) or excel_team.strip()
    norm_surname = _norm(surname)
    norm_team = _norm(canonical_team)

    # 1. Overrides file (highest priority — exact match on raw Excel values)
    ov_key = (_norm(excel_name), _norm(excel_team))
    if ov_key in overrides and overrides[ov_key]:
        return overrides[ov_key], "override"

    # 2. (surname, team) — the main matching path
    code = player_lookup.get((norm_surname, norm_team), "")
    if code:
        return code, "surname+team"

    # 3. Full Excel name (before initial stripping) + team
    code = player_lookup.get((_norm(excel_name.strip()), norm_team), "")
    if code:
        return code, "fullname+team"

    # 4. Unique surname anywhere across the whole season (ignores team)
    matches = surname_all.get(norm_surname, [])
    if len(matches) == 1:
        return matches[0], "unique surname"

    return "", ""


def _resolve_gk_team(
    excel_name: str,
    team_id_lookup: dict[str, str],
    season_id: str,
    overrides: dict[tuple[str, str], str],
) -> tuple[str, str]:
    """
    Match GK team name to a player_code.
    Returns (player_code, method_description) or ("", "") if unmatched.
    """
    # Override: key = (norm_excel_name, "")
    ov_key = (_norm(excel_name), "")
    if ov_key in overrides and overrides[ov_key]:
        return overrides[ov_key], "override"

    canonical = _normalize_team(excel_name) or excel_name.strip()
    team_id = team_id_lookup.get(canonical, "")
    if team_id:
        return f"{season_id}-team-{team_id}", "team alias"

    # Try raw name directly against full_name lookup
    team_id = team_id_lookup.get(excel_name.strip(), "")
    if team_id:
        return f"{season_id}-team-{team_id}", "exact team name"

    return "", ""


# ---------------------------------------------------------------------------
# Config loaders
# ---------------------------------------------------------------------------


def _load_overrides(overrides_path: str | None) -> dict[tuple[str, str], str]:
    """
    Load (norm_excel_name, norm_excel_team) -> player_code from a CSV file.
    For GK teams, excel_team should be blank in the CSV.
    """
    if not overrides_path:
        return {}
    path = Path(overrides_path)
    if not path.exists():
        # Not an error: the file only exists once a run has produced unmatched
        # rows. Carry on; if anything is unmatched the run will write it here.
        print(f"NOTE: no overrides file at {path} yet — continuing without it.")
        return {}
    df = pd.read_csv(path, dtype=str).fillna("")
    result: dict[tuple[str, str], str] = {}
    for _, r in df.iterrows():
        key = (_norm(str(r.get("excel_name", ""))), _norm(str(r.get("excel_team", ""))))
        code = str(r.get("player_code", "")).strip()
        if code:
            result[key] = code
    return result


def _find_import_file() -> Path:
    """Return the single Excel workbook sitting in the import/ drop folder."""
    drop = f"Drop this season's FOOTY20XX.xlsx in '{IMPORT_DIR}/' and re-run."
    if not IMPORT_DIR.exists():
        print(f"ERROR: no '{IMPORT_DIR}/' folder. Create it, then: {drop}")
        sys.exit(1)

    candidates = sorted(
        p
        for p in IMPORT_DIR.iterdir()
        if p.is_file() and p.suffix.lower() in {".xlsx", ".xls"} and not p.name.startswith("~$")
    )
    if not candidates:
        print(f"ERROR: no .xlsx workbook in '{IMPORT_DIR}/'. {drop}")
        sys.exit(1)
    if len(candidates) > 1:
        print(f"ERROR: more than one workbook in '{IMPORT_DIR}/' - leave just one:")
        for p in candidates:
            print(f"  {p.name}")
        sys.exit(1)

    chosen = candidates[0]
    if chosen.suffix.lower() == ".xls":
        print(f"ERROR: '{chosen.name}' is an old .xls - save it as .xlsx in Excel, then re-run.")
        sys.exit(1)
    return chosen


def _detect_season(players_df: pd.DataFrame) -> str:
    """Return the most recent season in players.csv by lexicographic sort."""
    seasons = [s for s in players_df["season"].unique() if s]
    if not seasons:
        print("ERROR: players.csv is empty. Run 'uv run sync-scores' first.")
        sys.exit(1)
    return sorted(seasons)[-1]


def _warn_if_stale(season_id: str) -> None:
    """Nudge to re-run sync-scores if seasons.csv shows players.csv is old / unsynced."""
    seasons_csv = DATA_DIR / "seasons.csv"
    if not seasons_csv.exists():
        return
    df = pd.read_csv(seasons_csv, dtype=str).fillna("")
    row = df[df["season_id"] == season_id]
    ts = row.iloc[0].get("last_synced_at", "") if not row.empty else ""
    if not ts:
        print(f"NOTE: no last_synced_at for {season_id} — run 'uv run sync-scores' first.")
        return
    try:
        synced = datetime.fromisoformat(ts)
    except ValueError:
        return
    age = datetime.now(UTC) - synced
    if age > timedelta(days=STALE_SYNC_DAYS):
        print(
            f"NOTE: players.csv for {season_id} last synced {age.days}d ago ({synced.date()}). "
            "Run 'uv run sync-scores' for up-to-date squads/prices."
        )


# ---------------------------------------------------------------------------
# Main import logic
# ---------------------------------------------------------------------------

def _suggestion_index(
    players_df: pd.DataFrame, season_id: str
) -> tuple[list[tuple[str, str, str, str]], list[tuple[str, str]]]:
    """
    Build lookup data for fixes-file suggestions and --auto.

    Returns (players, teams):
      players = [(norm_name_part, "Full Name (Club)", player_code, norm_club), ...]
                one row per significant word of the player's full/friendly name
      teams   = [(norm_team_name, "Team Name -> team_code"), ...]
    """
    sp = players_df[players_df["season"] == season_id]
    tid_to_name = {
        str(r["element_id"]): str(r["full_name"]) for _, r in sp[sp["type"] == "team"].iterrows()
    }

    players: list[tuple[str, str, str, str]] = []
    for _, r in sp[sp["type"] == "player"].iterrows():
        club = tid_to_name.get(str(r.get("team_id", "")), "")
        code = str(r["code"])
        name = str(r.get("full_name", "") or r.get("friendly_name", "") or "")
        disp = f"{name} ({club})" if club else name
        keys: set[str] = set()
        for src in (str(r.get("full_name", "") or ""), str(r.get("friendly_name", "") or "")):
            words = src.split()
            for w in words:
                nw = _norm(w)
                if len(nw) >= 3:
                    keys.add(nw)
            if words:  # keep the last word even if short (e.g. "Son")
                keys.add(_norm(words[-1]))
        for key in keys:
            if key:
                players.append((key, disp, code, _norm(club)))

    teams = [
        (_norm(name), f"{name} -> {season_id}-team-{eid}") for eid, name in tid_to_name.items()
    ]
    return players, teams


def _same_club(a: str, b: str) -> bool:
    """Loose club match — tolerates 'Ipswich' vs 'Ipswich Town', etc."""
    return bool(a) and bool(b) and (a == b or a in b or b in a)


def _suggest(
    u: dict,
    players: list[tuple[str, str, str, str]],
    teams: list[tuple[str, str]],
) -> str:
    """Return up to 3 'player_code  Full Name (Club)' guesses for one unmatched row."""
    def ratio(a: str, b: str) -> float:
        return difflib.SequenceMatcher(None, a, b).ratio()

    if u["position"] == "GK":
        target = _norm(_normalize_team(u["excel_name"]) or u["excel_name"])
        hits = [disp for key, disp in teams if ratio(target, key) >= 0.6]
        return " | ".join(hits[:3])

    target = _norm(_extract_surname(u["excel_name"]))
    club = _norm(_normalize_team(u["excel_team"]) or u["excel_team"])
    scored: list[tuple[float, str]] = []
    for key, disp, code, cclub in players:
        score = ratio(target, key)
        if _same_club(club, cclub):
            score += 0.15  # same club — nudge it up
        if score >= 0.6:
            scored.append((score, f"{code}  {disp}"))
    scored.sort(key=lambda s: s[0], reverse=True)
    out: list[str] = []
    for _, text in scored:
        if text not in out:
            out.append(text)
        if len(out) == 3:
            break
    return " | ".join(out)


def _auto_pick(
    excel_name: str, excel_team: str, players: list[tuple[str, str, str, str]]
) -> tuple[str, float]:
    """
    --auto: return (player_code, score) when one same-club candidate is a confident
    surname match and clearly ahead of the rest, else ("", 0.0).
    """
    target = _norm(_extract_surname(excel_name))
    club = _norm(_normalize_team(excel_team) or excel_team)
    if not (target and club):
        return "", 0.0

    best: dict[str, float] = {}
    for key, _disp, code, cclub in players:
        if not _same_club(club, cclub):
            continue
        score = difflib.SequenceMatcher(None, target, key).ratio()
        best[code] = max(best.get(code, 0.0), score)

    ranked = sorted(((s, c) for c, s in best.items() if s >= AUTO_MIN_SCORE), reverse=True)
    if len(ranked) == 1:
        return ranked[0][1], ranked[0][0]
    if len(ranked) >= 2 and ranked[0][0] - ranked[1][0] >= 0.10:
        return ranked[0][1], ranked[0][0]
    return "", 0.0


def _write_fixes_file(
    to_fix: list[dict],
    path: Path,
    players: list[tuple[str, str, str, str]],
    teams: list[tuple[str, str]],
) -> None:
    """
    Write rows needing a manual code to a fixes CSV. Each row carries a 'note'
    (why it's here) and a 'player_code' that is blank for an unmatched row or
    pre-filled with the current guess for a suspect (club mismatch / duplicate)
    row — leave it to accept, change it to correct.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(
        [
            {
                "manager": r["manager"],
                "position": r["position"],
                "excel_name": r["excel_name"],
                "excel_team": r["excel_team"],
                "player_code": r.get("player_code", ""),
                "note": r.get("note", ""),
                "suggestions": _suggest(r, players, teams),
            }
            for r in to_fix
        ]
    ).to_csv(path, index=False)


def _import_workbook(
    xlsx_path: Path,
    season_id: str,
    players_df: pd.DataFrame,
    overrides: dict[tuple[str, str], str],
    overrides_path: str | None,
    shape: tuple[int, int, int],
    budget: int,
    strict: bool,
    auto: bool,
    dry_run: bool,
) -> bool:
    """Import the workbook. Returns True only if rows were written to disk."""
    team_id_lookup, tid_to_name, player_lookup, surname_all = _build_lookups(players_df, season_id)

    if not team_id_lookup:
        print(f"ERROR: no team data for season {season_id} in players.csv.")
        print("Run 'uv run sync-scores' to populate it, then retry.")
        sys.exit(1)

    sugg_players, sugg_teams = _suggestion_index(players_df, season_id)
    position_order = _position_order(shape)
    n_rows = len(position_order)  # 11 = 1 GK + 10 outfield

    _sp = players_df[players_df["season"] == season_id]
    code_to_name = dict(zip(_sp["code"].astype(str), _sp["full_name"].astype(str)))
    _pl = _sp[_sp["type"] == "player"]
    code_to_club = {
        c: tid_to_name.get(t, "")
        for c, t in zip(_pl["code"].astype(str), _pl["team_id"].astype(str))
    }

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    all_rows: list[dict] = []
    unmatched: list[dict] = []
    review: list[dict] = []        # matched but suspect (club mismatch / duplicate) -> fixes.csv
    auto_matched: list[dict] = []
    warnings: list[str] = []       # soft — block only under --strict
    blockers: list[str] = []       # hard — always block a real import
    accepted_club: list[str] = []  # club mismatch OK'd by being pinned in fixes.csv
    underspent: list[str] = []

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in SKIP_SHEETS:
            continue

        manager = canonicalize(sheet_name)
        ws = wb[sheet_name]
        sheet_rows = list(ws.iter_rows(min_row=8, max_row=8 + n_rows - 1, values_only=True))
        declared_total = ws.cell(row=8 + n_rows, column=3).value  # row 19: sheet's own squad total

        if len(sheet_rows) < n_rows:
            print(f"WARN: '{sheet_name}' has {len(sheet_rows)} rows, expected {n_rows} — skipping")
            continue

        print(f"\n=== {manager} ===")
        sheet_cost = 0

        for raw_row, position in zip(sheet_rows[:n_rows], position_order):
            col_a = str(raw_row[0] or "").strip()
            col_b = str(raw_row[1] or "").strip() if len(raw_row) > 1 else ""
            col_c = raw_row[2] if len(raw_row) > 2 else None
            cost = _clean_cost(col_c)

            if not col_a:
                print(f"  SKIP  [{position}] empty cell")
                continue
            sheet_cost += int(cost)

            if position == "GK":
                player_code, method = _resolve_gk_team(col_a, team_id_lookup, season_id, overrides)
                if player_code:
                    print(f"  GK    {col_a:<25s} -> {player_code}  £{cost}  [{method}]")
                else:
                    print(f"  UNMATCHED GK: '{col_a}' — not found in players.csv for {season_id}")
                    unmatched.append({"manager": manager, "position": position, "excel_name": col_a, "excel_team": ""})
                    continue
            else:
                player_code, method = _match_player(col_a, col_b, player_lookup, surname_all, overrides)
                if not player_code and auto:
                    player_code, score = _auto_pick(col_a, col_b, sugg_players)
                    if player_code:
                        method = f"auto {score:.2f}"
                        auto_matched.append(
                            {"manager": manager, "position": position, "excel_name": col_a,
                             "excel_team": col_b, "player_code": player_code, "score": score}
                        )
                if player_code:
                    print(f"  {position:<5s} {col_a:<25s} {col_b:<20s} -> {player_code}  £{cost}  [{method}]")
                    got_club = code_to_club.get(player_code, "")
                    want_club = _normalize_team(col_b) or col_b
                    if col_b and got_club and not _same_club(_norm(want_club), _norm(got_club)):
                        pname = code_to_name.get(player_code, "")
                        detail = f"{manager} {col_a}: '{col_b}' -> {got_club} ({pname})"
                        if method == "override":
                            # code pinned in fixes.csv = "I've checked this" (e.g. transfer)
                            accepted_club.append(detail)
                        else:
                            blockers.append(f"{detail} -- sheet club doesn't match {player_code}")
                            review.append({
                                "manager": manager, "position": position,
                                "excel_name": col_a, "excel_team": col_b,
                                "player_code": player_code,  # keep to accept, change to fix
                                "note": f"club? matched {got_club}, sheet says '{col_b}'",
                            })
                else:
                    print(f"  UNMATCHED {position}: '{col_a}' @ '{col_b}'")
                    unmatched.append({"manager": manager, "position": position, "excel_name": col_a, "excel_team": col_b})
                    continue

            all_rows.append({
                "player_code": player_code,
                "season_id": season_id,
                "manager_name": manager,
                "position": position,
                "cost": cost,
                "gw_from": "1",
                "gw_to": "38",
                "excel_name": col_a,
                "excel_team": "" if position == "GK" else col_b,
            })

        declared = _clean_cost(declared_total) if declared_total not in (None, "") else ""
        if declared and int(declared) != sheet_cost:
            warnings.append(f"{manager}: sheet total £{declared} but costs add up to £{sheet_cost}")
        if sheet_cost > budget:
            blockers.append(f"{manager}: squad costs £{sheet_cost}, over the £{budget} limit")
        elif sheet_cost < budget:
            underspent.append(f"{manager} £{sheet_cost}")

    # ---- verification: duplicate picks (a player / GK team owned by >1 squad) ----
    counts: dict[str, int] = {}
    for r in all_rows:
        counts[r["player_code"]] = counts.get(r["player_code"], 0) + 1
    for code, n in sorted(counts.items()):
        if n <= 1:
            continue
        rows = [r for r in all_rows if r["player_code"] == code]
        who = sorted(f"{r['manager_name']} ({r['position']})" for r in rows)
        blockers.append(f"{code} {code_to_name.get(code, '')} picked {n}×: {', '.join(who)}")
        for r in rows:
            review.append({
                "manager": r["manager_name"], "position": r["position"],
                "excel_name": r["excel_name"], "excel_team": r["excel_team"],
                "player_code": r["player_code"], "note": f"duplicate {n}× — change one",
            })
    if warnings and strict:
        blockers.append(f"--strict: {len(warnings)} warning(s) listed above")

    # one fixes row per (manager, name, team); keep the pre-filled code, merge notes
    merged: dict[tuple[str, str, str], dict] = {}
    for r in review:
        k = (r["manager"], r["excel_name"], r["excel_team"])
        if k in merged:
            merged[k]["note"] += f"; {r['note']}"
        else:
            merged[k] = dict(r)
    to_fix = [{**u, "player_code": "", "note": "unmatched"} for u in unmatched]
    to_fix += list(merged.values())

    fixes_path = Path(overrides_path) if overrides_path else IMPORT_DIR / "fixes.csv"
    fix_hint = ""
    if to_fix:
        if fixes_path.exists():
            tail = "." if overrides_path else f" with --overrides {fixes_path}."
            fix_hint = f"{fixes_path} kept as-is — update it and re-run{tail}"
        else:
            _write_fixes_file(to_fix, fixes_path, sugg_players, sugg_teams)
            fix_hint = (
                f"Wrote {len(to_fix)} row(s) to {fixes_path}. Read 'note', set/correct "
                f"'player_code' (see 'suggestions'), then re-run with --overrides {fixes_path}."
            )

    # Summary
    print(f"\n{'=' * 60}")
    print(f"Season:    {season_id}")
    print(f"Matched:   {len(all_rows)}" + (f" ({len(auto_matched)} auto)" if auto_matched else ""))
    print(f"Unmatched: {len(unmatched)}")

    if auto_matched:
        print("\nAuto-matched — CHECK THESE, they are guesses:")
        for a in auto_matched:
            name = code_to_name.get(a["player_code"], "")
            print(
                f"  {a['excel_name']} @ {a['excel_team']}  ->  {a['player_code']}  {name}"
                f"  [{a['score']:.2f}]  ({a['manager']} / {a['position']})"
            )

    if accepted_club:
        print(f"\nClub mismatch OK'd — code pinned in fixes.csv ({len(accepted_club)}):")
        for a in accepted_club:
            print(f"  - {a}")

    if underspent:
        print(f"\nUnderspent (below £{budget}, allowed): " + ", ".join(underspent))

    if warnings:
        print(f"\nWARNINGS ({len(warnings)}):")
        for w in warnings:
            print(f"  - {w}")

    if blockers:
        print(f"\nPROBLEMS ({len(blockers)}) — these block a real import:")
        for b in blockers:
            print(f"  - {b}")

    if unmatched:
        print(f"\nUnmatched ({len(unmatched)}):")
        for u in unmatched:
            print(f"  {u['manager']:<16s} {u['position']:<4s} {u['excel_name']} @ {u['excel_team']}")

    if unmatched or blockers:
        verb = "Would refuse" if dry_run else "Refusing"
        print(
            f"\n{verb} to import: {len(unmatched)} unmatched, {len(blockers)} problem(s). "
            "Nothing was written."
        )
        if fix_hint:
            print(fix_hint)
        print("For a whole club that's mis-spelled, add it to scripts/team_aliases.py.")
        if unmatched and not auto:
            print("Or let --auto accept the confident guesses for you.")
        if not dry_run:
            sys.exit(1)
        return False

    if dry_run:
        print("\nDry run — all checks passed. No changes written.")
        return False

    if not all_rows:
        print("\nNothing to write.")
        return False

    csv_cols = ["player_code", "season_id", "manager_name", "position", "cost", "gw_from", "gw_to"]
    new_df = pd.DataFrame(all_rows)[csv_cols]
    sel_path = DATA_DIR / "manager_selections.csv"
    if sel_path.exists() and sel_path.stat().st_size > 0:
        existing = pd.read_csv(sel_path, dtype=str).fillna("")
        existing = existing[existing["season_id"] != season_id]
        result = pd.concat([existing, new_df], ignore_index=True)
    else:
        result = new_df
    result.to_csv(sel_path, index=False)
    print(f"\nWritten {len(new_df)} rows to {sel_path}")
    return True


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import manager selections from a FOOTY20XX.xlsx workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "xlsx",
        nargs="?",
        help="Path to the Excel workbook (e.g. FOOTY2026.xlsx). "
        "Omit to load the single workbook sitting in import/ (deleted after a successful import).",
    )
    parser.add_argument(
        "--season",
        help="Season ID to import into (e.g. 2026-27). Auto-detected from players.csv if omitted.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview matching only — no changes written to manager_selections.csv.",
    )
    parser.add_argument(
        "--overrides",
        metavar="FILE",
        help="CSV of manual name->player_code fixes. Fine to point at import/fixes.csv "
        "before it exists — the run creates it when there are unmatched players.",
    )
    parser.add_argument(
        "--auto",
        action="store_true",
        help="Auto-accept a confident same-club guess for an unmatched player (shown for review).",
    )
    parser.add_argument(
        "--shape",
        metavar="D-M-F",
        help="Outfield shape, e.g. 3-5-2. Default: per-season table in this script, else 3-4-3.",
    )
    parser.add_argument(
        "--budget",
        type=int,
        default=DEFAULT_BUDGET,
        help=f"Auction spend per manager (default {DEFAULT_BUDGET}); a mismatch warns.",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Treat budget / squad-total warnings as errors (blocks the import).",
    )
    args = parser.parse_args()

    if args.xlsx:
        xlsx_path = Path(args.xlsx)
        from_import_dir = False
        if not xlsx_path.exists():
            print(f"ERROR: workbook not found: {xlsx_path}")
            sys.exit(1)
    else:
        xlsx_path = _find_import_file()
        from_import_dir = True

    players_csv = DATA_DIR / "players.csv"
    if not players_csv.exists():
        print(f"ERROR: {players_csv} not found. Run 'uv run sync-scores' first.")
        sys.exit(1)

    players_df = pd.read_csv(players_csv, dtype=str).fillna("")
    season_id = args.season or _detect_season(players_df)

    season_players = players_df[players_df["season"] == season_id]
    if season_players.empty:
        print(f"ERROR: no players found for season '{season_id}' in players.csv")
        print(f"Available seasons: {sorted(players_df['season'].unique())}")
        sys.exit(1)

    shape = _parse_shape(args.shape) if args.shape else SEASON_SHAPES.get(season_id, DEFAULT_SHAPE)
    if sum(shape) != 10:
        print(f"ERROR: shape {shape[0]}-{shape[1]}-{shape[2]} (DEF-MID-FWD) must add up to 10.")
        sys.exit(1)
    if args.shape:
        shape_src = "--shape"
    elif season_id in SEASON_SHAPES:
        shape_src = f"SEASON_SHAPES[{season_id!r}]"
    else:
        shape_src = "default"

    print(f"Season:   {season_id}")
    print(f"Workbook: {xlsx_path}")
    print(f"Squad:    1 GK + {shape[0]} DEF + {shape[1]} MID + {shape[2]} FWD  ({shape_src})")
    if args.dry_run:
        print("Mode:     dry-run (no writes)")
    if args.overrides:
        print(f"Overrides: {args.overrides}")
    _warn_if_stale(season_id)

    overrides = _load_overrides(args.overrides)

    wrote = _import_workbook(
        xlsx_path=xlsx_path,
        season_id=season_id,
        players_df=players_df,
        overrides=overrides,
        overrides_path=args.overrides,
        shape=shape,
        budget=args.budget,
        strict=args.strict,
        auto=args.auto,
        dry_run=args.dry_run,
    )

    if wrote and from_import_dir:
        xlsx_path.unlink()
        print(f"Deleted {xlsx_path} after a successful import.")

    # The default fixes file is scratch work — clear it once the import succeeds.
    # A fixes file at a custom --overrides path is left alone (like an explicit workbook path).
    default_fixes = IMPORT_DIR / "fixes.csv"
    if (
        wrote
        and args.overrides
        and default_fixes.exists()
        and Path(args.overrides).resolve() == default_fixes.resolve()
    ):
        default_fixes.unlink()
        print(f"Deleted {default_fixes} after a successful import.")


if __name__ == "__main__":
    main()
